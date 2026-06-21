#!/usr/bin/env python3
"""Export baseline reproduction CSV into an Excel workbook."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
CSV_PATH = ROOT / "reproduction_results.csv"
XLSX_PATH = ROOT / "reproduction_results.xlsx"


def read_rows():
    with CSV_PATH.open(newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def add_long_table(wb: Workbook, rows: list[dict[str, str]]):
    ws = wb.create_sheet("long_results")
    ws.title = "long_results"
    headers = list(rows[0].keys()) if rows else []
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    style_sheet(ws)
    ws.freeze_panes = "A2"


def add_matrix(wb: Workbook, rows: list[dict[str, str]]):
    ws = wb.active
    ws.title = "matrix"
    datasets = [
        "Cora",
        "CiteSeer",
        "PubMed",
        "Amazon-Photo",
        "Amazon-Computers",
        "Coauthor-CS",
        "Coauthor-Physics",
        "Wiki-CS",
        "Actor",
        "Chameleon",
        "Squirrel",
        "Cornell",
        "Texas",
        "Wisconsin",
    ]
    baselines = []
    for row in rows:
        baseline = row["baseline"]
        if baseline not in baselines:
            baselines.append(baseline)

    ws.append(["Baseline", *datasets])
    by_key = {(r["baseline"], r["dataset"]): r for r in rows}
    for baseline in baselines:
        out = [baseline]
        for dataset in datasets:
            row = by_key.get((baseline, dataset))
            if not row:
                out.append("")
                continue
            status = row.get("status", "")
            mean = row.get("reproduced_mean", "")
            std = row.get("reproduced_std", "")
            metric = row.get("metric", "")
            runs = row.get("runs", "")
            if mean:
                val = f"{metric} {float(mean) * 100:.2f}"
                if std:
                    val += f"±{float(std) * 100:.2f}"
                val += f"; {runs}"
            else:
                val = status
            out.append(val)
        ws.append(out)
    style_sheet(ws)
    ws.freeze_panes = "B2"


def add_protocol(wb: Workbook):
    ws = wb.create_sheet("protocol")
    protocol_path = ROOT / "reproduction_protocol.md"
    if protocol_path.exists():
        for line in protocol_path.read_text(encoding="utf-8").splitlines():
            ws.append([line])
    style_sheet(ws)


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, max_len + 2)


def main():
    rows = read_rows()
    wb = Workbook()
    add_matrix(wb, rows)
    add_long_table(wb, rows)
    add_protocol(wb)
    wb.save(XLSX_PATH)
    print(XLSX_PATH)


if __name__ == "__main__":
    main()
